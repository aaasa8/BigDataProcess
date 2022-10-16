#!/usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']
score = []
grade = []
row_id = 1;
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		score.append(sum_v)
	row_id += 1
score.reverse()

for i in range(0, len(score)):
	if i <= len(score) * 0.3:
		if i <= len(score) * 0.3 * 0.5:
			grade.append("A+")
		else:
			grade.append("A0")
	elif i <= len(score) * 0.7:
		if i <= len(score) * 0.7 * 0.5:
			grade.append("B+")
		else:
			grade.append("B0")
	else:
		if i <= len(score) * 0.7 + len(score) * 0.7 / 4:
			grade.append("C+")
		else:
			grade.append("C0")
score_grade = dict(zip(score, grade))

row_id = 1
for row in ws:
	if row_id != 1:
		c = ws.cell(row = row_id, column = 7).value
		ws.cell(row = row_id, column = 8).value = score_grade[c]
	row_id += 1

wb.save("student.xlsx")
